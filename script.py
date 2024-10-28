from openpyxl import load_workbook

def extract_companies(filename):
  workbook = load_workbook(filename)
  sheet = workbook.active

  companies = []
  for row in sheet.iter_rows(min_row=2, values_only=True):
    company_name, phone_number = row
    if company_name not in [company[0] for company in companies]:
      companies.append([company_name, phone_number])

  return companies

if __name__ == "__main__":
  filename = "your_excel_file.xlsx"  # Replace with your file name
  companies = extract_companies(filename)

  for company in companies:
    print(f"Company: {company[0]}, Phone: {company[1]}")
